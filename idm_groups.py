# This script will extract both group information and their corresponding members from Red Hat IDM. It will then compile this data into an organized groups.xlsx file, where each sheet will represent a group along with its respective members.
# Author: Andre Facina
# Tested in Red Hat Identity Management version: 4.10.1
# Usage: python3 idm_groups.py  --server IDM_IP_OR_HOSTNAME --bind_dn "cn=Directory Manager"

import ldap3
from openpyxl import Workbook
import re
import argparse
import getpass

def main():
    parser = argparse.ArgumentParser(description="LDAP Group Extraction Tool")
    parser.add_argument("--server", required=True, help="LDAP server URI")
    parser.add_argument("--bind_dn", required=True, help="Bind DN")
    parser.add_argument("--bind_password", help="Bind password. You can omit it and avoid clear text")

    args = parser.parse_args()

    if not args.bind_password:
        args.bind_password = getpass.getpass("Enter bind password: ")

    # Connection settings
    server = args.server
    bind_dn = args.bind_dn
    bind_password = args.bind_password

    base_dn = 'dc=example,dc=redhat,dc=com'
    group_search_filter = '(objectClass=groupofnames)'
    group_attributes = ['cn', 'member']

    # Connection
    server = ldap3.Server(server)
    connection = ldap3.Connection(server, user=bind_dn, password=bind_password)

    if not connection.bind():
        print("Failed to bind to the server:", connection.result)
    else:
        print("Connected to the server")

# Extract members of group
        groups = []
        connection.search(base_dn, group_search_filter, attributes=group_attributes)
        if connection.response:
            for entry in connection.response:
                groups.append(entry['attributes'])

        workbook = Workbook()

# Write the group and members in the xlsx
        for group in groups:
            group_name = group.get('cn', ['N/A'])[0]
            sanitized_group_name = re.sub(r'[\/:*?"<>|]', '_', group_name)
            sheet = workbook.create_sheet(sanitized_group_name)

            members = group.get('member', [])
            for index, member in enumerate(members, start=1):
                sheet.cell(row=index, column=1, value=member)

        workbook.save('groups.xlsx')
        print("Groups written to groups.xlsx")

        connection.unbind()

if __name__ == "__main__":
    main()

